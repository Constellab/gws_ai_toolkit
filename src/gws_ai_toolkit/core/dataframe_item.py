import uuid
from typing import Dict, List

from openpyxl import load_workbook
from pandas import DataFrame, ExcelFile, read_csv, read_excel
from pydantic import BaseModel


class DataFrameItemInfo(BaseModel):
    id: str
    name: str


class DataFrameItem:
    """Class to represent a dataframe with its name and file path"""

    id: str
    name: str
    file_path: str

    _cached_dataframes: Dict[str, DataFrame]  # Cache for sheet dataframes
    _sheet_names: List[str]  # Cache for sheet names

    def __init__(self, name: str, file_path: str):
        self.id = str(uuid.uuid4())  # Generate unique ID
        self.name = name
        self.file_path = file_path
        self._cached_dataframes = {}  # Cache for sheet dataframes
        self._sheet_names = []  # Cache for sheet names

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names for Excel files, or empty list for CSV"""
        if self._sheet_names is not None:
            return self._sheet_names

        ext = self.file_path.lower().split('.')[-1]
        if ext in ["xls", "xlsx"]:
            excel_file = ExcelFile(self.file_path)
            self._sheet_names = excel_file.sheet_names
        else:
            self._sheet_names = []

        return self._sheet_names

    def has_multiple_sheets(self) -> bool:
        """Check if the file has multiple sheets"""
        return len(self.get_sheet_names()) > 1

    def get_default_dataframe(self) -> DataFrame:
        """Get the dataframe of the first sheet (or CSV data)"""
        ext = self.file_path.lower().split('.')[-1]

        if ext == "csv":
            return self.get_dataframe("")
        elif ext in ["xls", "xlsx"]:
            sheet_names = self.get_sheet_names()
            if sheet_names:
                return self.get_dataframe(sheet_names[0])

        return DataFrame()

    def _handle_merged_cells(self, file_path: str, sheet_name: str = None) -> DataFrame:
        """Handle merged cells in Excel files by filling all cells in merged ranges"""
        # Load the workbook with openpyxl
        workbook = load_workbook(file_path, data_only=True)

        # Get the worksheet
        if sheet_name:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        # Get all merged cell ranges
        merged_ranges = worksheet.merged_cells.ranges

        # First, read the Excel file normally with pandas
        if sheet_name:
            df = read_excel(file_path, sheet_name=sheet_name)
        else:
            df = read_excel(file_path)

        # Process each merged range
        for merged_range in merged_ranges:
            # Get the value from the top-left cell of the merged range
            top_left_cell = worksheet[merged_range.coord.split(':')[0]]
            merged_value = top_left_cell.value

            # Convert openpyxl coordinates to pandas DataFrame indices
            min_row = merged_range.min_row - 2  # -1 for 0-based indexing, -1 for header row
            max_row = merged_range.max_row - 2
            min_col = merged_range.min_col - 1  # -1 for 0-based indexing
            max_col = merged_range.max_col - 1

            # Ensure indices are within DataFrame bounds
            min_row = max(0, min_row)
            max_row = min(len(df) - 1, max_row)
            min_col = max(0, min_col)
            max_col = min(len(df.columns) - 1, max_col)

            # Fill all cells in the merged range with the same value
            if min_row <= max_row and min_col <= max_col:
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row < len(df) and col < len(df.columns):
                            df.iloc[row, col] = merged_value

        return df

    def get_dataframe(self, sheet_name: str = "") -> DataFrame:
        """Get dataframe for specific sheet (or CSV data if sheet_name is empty)"""
        # Use cached dataframe if available
        cache_key = sheet_name if sheet_name else "default"
        if cache_key in self._cached_dataframes:
            return self._cached_dataframes[cache_key]

        ext = self.file_path.lower().split('.')[-1]

        if ext == "csv":
            # Try to detect separator automatically
            with open(self.file_path, 'r', encoding='utf-8') as f:
                sample = f.read(1024)
            if ';' in sample and sample.count(';') > sample.count(','):
                sep = ';'
            else:
                sep = ','
            dataframe = read_csv(self.file_path, sep=sep)
        elif ext in ["xls", "xlsx"]:
            if sheet_name:
                dataframe = self._handle_merged_cells(self.file_path, sheet_name=sheet_name)
            else:
                # Use first sheet if no sheet specified
                sheet_names = self.get_sheet_names()
                if sheet_names:
                    dataframe = self._handle_merged_cells(self.file_path, sheet_name=sheet_names[0])
                else:
                    dataframe = DataFrame()
        else:
            dataframe = DataFrame()

        # Clean up the dataframe
        dataframe = dataframe.dropna(axis=1, how='all')

        # Cache the dataframe
        self._cached_dataframes[cache_key] = dataframe

        return dataframe

    def to_info(self) -> DataFrameItemInfo:
        """Convert to DataFrameInfo for API compatibility"""
        return DataFrameItemInfo(id=self.id, name=self.name)
